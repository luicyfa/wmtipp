from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path("/Users/bjornbochinski/Downloads/WCup_2026_4.2.6_de.xlsx")
ROOT = Path(__file__).resolve().parents[1]
SEED = ROOT / "db" / "seed.sql"
CSV_OUT = ROOT / "db" / "worldcup_2026_matches.csv"

PLAYER_HASHES = {
    "Admin": "scrypt:af51c02a34f7ac1f6095de561ae78e8d:ae6006e3b6a1eeb4b88cb1dc54ce077d199d3d5d3650092b0f9fb43b41dc6c1169efeb3681bd52ff0e1c7092216b71129ed9397f70e444f83aeb01e4ba2a8ed6",
    "Björn": "scrypt:67254c5b9a9c0d7afc1a9336b423857f:9344f72a209abc1389eb504b9b38820cf70ab3d421f368c9b222d6dccdee39298b3cde437ff5ebd36ab473abe82be0d6de915bd3c7739b562be45866bc6a813d",
    "Henry": "scrypt:950470ff4d40b6aa13ac3422aab9e80d:d2b6e3780ad759509581f17f6819fb70f0e6147e27daaacee9e0c83f735c9dd5f776b32fb8b178269467dcf0694855587b104fcf6938c5a34f69bb01b6683bd8",
    "Opa": "scrypt:f8363f223534c54b3cbddb16ca4b2b4b:0dfe9ee8b64b817b1cba78479b7cc4044a85a2f6ab24a7af63831cf3ec972b929248a5297875f31ec5b676b956ce1a7220419e7badca147b9f64bdded21c68c8",
    "Oma": "scrypt:cfdda54687c64ec905ca66a341d5d9ba:36e04249f92ed7be21002cd2a5d07b6654b8d20a4b68bb48b4bda42c6c9fbf08352cb6da327972370d1d4212b77b58c00fa421585d2f7afea6d1470d79fa73b9",
}


def sql(value):
    if value is None or value == "":
        return "null"
    if isinstance(value, int):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def round_for(match_number: int, teams_code: str) -> tuple[str, str | None]:
    if match_number <= 72:
        group = teams_code[0] if teams_code and teams_code[0].isalpha() else None
        return "Gruppenphase", group
    if match_number <= 88:
        return "Sechzehntelfinale", None
    if match_number <= 96:
        return "Achtelfinale", None
    if match_number <= 100:
        return "Viertelfinale", None
    if match_number <= 102:
        return "Halbfinale", None
    if match_number == 103:
        return "Spiel um Platz 3", None
    return "Finale", None


def main():
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    teams: list[tuple[str, str, str]] = []
    groups = wb["Groups"]
    current_group = None
    for row in groups.iter_rows(min_row=6, values_only=True):
        code, _, name = row[1], row[2], row[3]
        if isinstance(name, str) and len(name) == 1 and name.isalpha():
            current_group = name
            continue
        if isinstance(code, str) and isinstance(name, str) and code[:1].isalpha():
            teams.append((name, code, current_group or code[:1]))

    matches = []
    ws = wb["Matches"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        match_number = row[1]
        if not isinstance(match_number, int):
            continue
        team_code_home, team_code_away = row[2], row[3]
        kickoff = row[5] if isinstance(row[5], datetime) else row[4]
        venue = row[7]
        home_label = row[8] or team_code_home
        away_label = row[9] or team_code_away
        round_name, group_code = round_for(match_number, str(team_code_home or ""))
        matches.append(
            {
                "match_number": match_number,
                "round": round_name,
                "group_code": group_code,
                "home_team_label": home_label,
                "away_team_label": away_label,
                "kickoff_at": kickoff.isoformat() + "+02:00",
                "venue": venue,
            }
        )

    with CSV_OUT.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(matches[0].keys()))
        writer.writeheader()
        writer.writerows(matches)

    lines = [
        "-- Seed data generated from WCup_2026_4.2.6_de.xlsx.",
        "-- Default PINs: Admin 1234, Björn 1111, Henry 2222, Opa 3333, Oma 4444.",
        "",
        "insert into score_rules (exact_score_points, tendency_points, goal_difference_points, team_goals_points)",
        "select 5, 3, 1, 1 where not exists (select 1 from score_rules);",
        "",
    ]
    for name, pin_hash in PLAYER_HASHES.items():
        lines.append(
            "insert into players (name, pin_hash, is_admin) values "
            f"({sql(name)}, {sql(pin_hash)}, {'true' if name == 'Admin' else 'false'}) "
            "on conflict (lower(name)) do update set pin_hash = excluded.pin_hash, is_admin = excluded.is_admin;"
        )

    lines.append("")
    for name, code, group in teams:
        short = code
        lines.append(
            "insert into teams (name, short_name, group_code, placeholder) values "
            f"({sql(name)}, {sql(short)}, {sql(group)}, false) "
            "on conflict (lower(name)) do update set short_name = excluded.short_name, group_code = excluded.group_code;"
        )

    lines.append("")
    for match in matches:
        lines.append(
            "insert into matches (match_number, round, group_code, home_team_label, away_team_label, kickoff_at, venue, status) values "
            f"({match['match_number']}, {sql(match['round'])}, {sql(match['group_code'])}, "
            f"{sql(match['home_team_label'])}, {sql(match['away_team_label'])}, "
            f"{sql(match['kickoff_at'])}, {sql(match['venue'])}, 'scheduled') "
            "on conflict (match_number) do update set "
            "round = excluded.round, group_code = excluded.group_code, home_team_label = excluded.home_team_label, "
            "away_team_label = excluded.away_team_label, kickoff_at = excluded.kickoff_at, venue = excluded.venue;"
        )

    SEED.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {len(teams)} teams and {len(matches)} matches")


if __name__ == "__main__":
    main()
